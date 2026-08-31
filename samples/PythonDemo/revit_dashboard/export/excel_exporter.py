"""Export filtered dashboard data to Excel."""

from datetime import datetime

from openpyxl import Workbook

from revit_dashboard.analytics.engine import apply_filters, build_dataframe, build_kpis
from revit_dashboard.contracts.payload import DashboardFilterState


def _ask_save_path(default_name: str) -> str | None:
    """Open a Windows Save File dialog and return the chosen path, or None if cancelled."""
    try:
        from System.Windows.Forms import SaveFileDialog, DialogResult

        dialog = SaveFileDialog()
        dialog.Title = "Export Dashboard Data"
        dialog.Filter = "Excel Files (*.xlsx)|*.xlsx|All Files (*.*)|*.*"
        dialog.FileName = default_name
        dialog.DefaultExt = ".xlsx"
        dialog.AddExtension = True

        if dialog.ShowDialog() == DialogResult.OK:
            return dialog.FileName
        return None
    except Exception as exc:
        print(f"[Export] SaveFileDialog failed ({exc}), using default path")
        return None


def export_filtered_rows_to_excel(
    all_rows: list[dict],
    filters: DashboardFilterState,
) -> str | None:
    """Write an ``.xlsx`` workbook with Summary, Filters, and Elements sheets.

    Opens a Windows Save File dialog for the user to choose where to save.
    Returns the file path on success, or None if the user cancelled.
    """
    df = build_dataframe(all_rows)
    filtered_df = apply_filters(df, filters)
    kpis = build_kpis(filtered_df)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    default_name = f"bim_dashboard_export_{timestamp}.xlsx"

    target = _ask_save_path(default_name)
    if target is None:
        return None

    wb = Workbook()

    # Summary sheet
    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_summary.append(["Metric", "Value"])
    for key, value in kpis.items():
        ws_summary.append([key, value])

    # Filters sheet
    ws_filters = wb.create_sheet("Filters")
    ws_filters.append(["Filter", "Values"])
    for key, value in filters.items():
        if isinstance(value, list):
            ws_filters.append([key, ", ".join(str(v) for v in value)])
        else:
            ws_filters.append([key, str(value)])

    # Elements sheet
    ws_data = wb.create_sheet("Elements")
    ws_data.append(filtered_df.columns)
    for row in filtered_df.iter_rows(named=False):
        ws_data.append(list(row))

    wb.save(target)
    return str(target)
